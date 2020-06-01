import time
import requests
import openpyxl as opx


def get_wallets(path_to_file):
    workbook = opx.load_workbook(path_to_file)
    sheet = workbook.worksheets[0]
    result = []
    for row in sheet.rows:
        result.append(row[0].value)
    return result


def get_response(wallet):
    response = requests.get('https://blockchair.com/ru/search', params={
        'q': wallet,
    })

    if response.status_code != 200:
        time.sleep(1)
        response = requests.get('https://blockchair.com/ru/search', params={
            'q': wallet,
        })
    elif response.status_code != 200:
        raise SystemError

    return response


def write_to_file(data, summa, path_to_file):
    wb = opx.Workbook()
    ws = wb.create_sheet('Биткоины', 0)
    ws.cell(row=1, column=1).value = 'Номер кошелька'
    ws.cell(row=1, column=2).value = 'Количество битков'

    i = 2
    for k, v in data.items():
        ws.cell(row=i, column=1).value = k
        ws.cell(row=i, column=2).value = v
        i += 1

    ws.cell(row=i, column=1).value = 'Итого'
    ws.cell(row=i, column=2).value = summa
    wb.save(path_to_file)


def main(data_file):
    wallets = get_wallets(data_file)
    wallets_with_coin = {}
    summa = 0

    for wallet in wallets:
        response = get_response(wallet)
        try:
            coin = response.text[response.text.find('Баланс: <span><span>') + 20:
                                 response.text.find('Баланс: <span><span>') + 100]
            coin_int = coin.split('<span')[0]
            coin_fract = coin.split('>')[1].split('</s')[0]
            coin_num = float(coin_int) + float(coin_fract)
        except IndexError:
            print('Нет такого кошелька')
            coin_num = 0
        # pprint(response.text)  # строка для дебаггинга
        # print(f'\n\n\n{coin}')  # строка для дебаггинга
        wallets_with_coin[wallet] = coin_num
        print(f'{wallet}: {coin_num} BTC')  # убрать строку
        summa += coin_num

    print(f'Сумма равна: {summa}\n')
    write_to_file(wallets_with_coin, summa, 'result.xlsx')


if __name__ == '__main__':
    main('wallets.xlsx')
